import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import pandas as pd

def criar_planilha_metricas():
    wb = openpyxl.Workbook()
    
    # Aba Resumo Geral
    ws = wb.active
    ws.title = "Resumo Geral"
    ws['A1'] = "Resumo das Métricas do Sistema"
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['Fila', 'População Média', 'Throughput', 'Utilização (%)', 'Tempo Médio de Resposta', 'Clientes Perdidos']
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
        ws.cell(row=2, column=col).font = Font(bold=True)
    
    # Dados das filas
    filas = ['Triagem', 'Emergência', 'Consulta', 'Internação', 'Alta']
    dados = [
        [0.55, 0.1000, 27.48, 0.00, 0],
        [0.92, 0.0304, 30.53, 0.14, 0],
        [1.57, 0.0696, 69.78, 2.44, 0],
        [6.85, 0.0056, 99.99, 1047.94, 7066],
        [0.28, 0.0732, 25.61, 0.36, 0]
    ]
    
    for row, (fila, dados_fila) in enumerate(zip(filas, dados), 3):
        ws.cell(row=row, column=1, value=fila)
        for col, valor in enumerate(dados_fila, 2):
            ws.cell(row=row, column=col, value=valor)
    
    # Aba Análise por Fila
    ws = wb.create_sheet("Análise por Fila")
    ws['A1'] = "Análise Detalhada por Fila"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Adicionar mais análises específicas por fila...
    
    # Aba Gráficos
    ws = wb.create_sheet("Gráficos")
    ws['A1'] = "Gráficos de Comparação"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Criar gráficos...
    
    wb.save('analise_metricas.xlsx')

def criar_planilha_comparacao():
    wb = openpyxl.Workbook()
    
    # Aba Comparação Original vs Melhorado
    ws = wb.active
    ws.title = "Comparação Original vs Melhorado"
    ws['A1'] = "Comparação entre Modelos Original e Melhorado"
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['Fila', 'Métrica', 'Original', 'Melhorado', 'Melhoria (%)']
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
        ws.cell(row=2, column=col).font = Font(bold=True)
    
    # Dados de comparação
    dados = [
        ['Triagem', 'População Média', 0.55, 0.45, 18.10],
        ['Triagem', 'Throughput', 0.1000, 0.1001, 0.07],
        ['Triagem', 'Utilização (%)', 27.48, 15.01, -12.48],
        ['Emergência', 'População Média', 0.92, 0.76, 17.23],
        ['Emergência', 'Throughput', 0.0304, 0.0304, -0.13],
        ['Emergência', 'Utilização (%)', 30.53, 19.04, -11.49],
        # Adicionar mais dados...
    ]
    
    for row, dados_linha in enumerate(dados, 3):
        for col, valor in enumerate(dados_linha, 1):
            ws.cell(row=row, column=col, value=valor)
    
    # Aba Análise de Impacto
    ws = wb.create_sheet("Análise de Impacto")
    ws['A1'] = "Análise de Impacto das Melhorias"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Adicionar análise de impacto...
    
    # Aba Recomendações
    ws = wb.create_sheet("Recomendações")
    ws['A1'] = "Recomendações e Próximos Passos"
    ws['A1'].font = Font(bold=True, size=14)
    
    recomendacoes = [
        "Implementar sistema de priorização de pacientes",
        "Considerar horários de pico para dimensionamento",
        "Avaliar a possibilidade de telemedicina para consultas não urgentes"
    ]
    
    for row, rec in enumerate(recomendacoes, 2):
        ws.cell(row=row, column=1, value=rec)
    
    wb.save('comparacao_modelos.xlsx')

if __name__ == "__main__":
    criar_planilha_metricas()
    criar_planilha_comparacao() 