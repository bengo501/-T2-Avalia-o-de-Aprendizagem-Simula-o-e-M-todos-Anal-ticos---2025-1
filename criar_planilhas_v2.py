import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def read_metrics_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    sections = content.split('##')
    data = {}
    
    # Parse Resumo Geral
    resumo_geral = {}
    for line in sections[1].split('\n'):
        if '|' in line and '---' not in line and 'Métrica' not in line:
            metric, value = [x.strip() for x in line.split('|')[1:3]]
            resumo_geral[metric] = value
    data['Resumo Geral'] = pd.DataFrame(list(resumo_geral.items()), columns=['Métrica', 'Valor'])
    
    # Parse Análise por Fila
    filas_data = {}
    current_fila = None
    for line in sections[2].split('\n'):
        if '###' in line:
            current_fila = line.replace('###', '').strip()
            filas_data[current_fila] = []
        elif '|' in line and '---' not in line and 'Métrica' not in line and current_fila:
            metric, value = [x.strip() for x in line.split('|')[1:3]]
            filas_data[current_fila].append([metric, value])
    
    for fila in filas_data:
        data[fila] = pd.DataFrame(filas_data[fila], columns=['Métrica', 'Valor'])
    
    return data

def read_comparison_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    sections = content.split('##')
    data = {}
    
    # Parse Comparação entre Modelos
    for section in sections:
        if 'COMPARAÇÃO ENTRE MODELOS' in section:
            tables = []
            current_table = []
            table_name = ''
            
            for line in section.split('\n'):
                if '|' in line:
                    if 'Métrica' in line:
                        if current_table:
                            tables.append((table_name, current_table))
                            current_table = []
                        table_name = line.split('|')[1].strip()
                    elif '---' not in line:
                        row = [x.strip() for x in line.split('|')[1:-1]]
                        if row:
                            current_table.append(row)
            
            if current_table:
                tables.append((table_name, current_table))
            
            for name, table in tables:
                df = pd.DataFrame(table[1:], columns=table[0])
                data[name] = df
    
    return data

def format_excel(wb, sheet):
    for column in sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    
    for row in sheet.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

def create_metrics_excel(data, output_file):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write Resumo Geral
        data['Resumo Geral'].to_excel(writer, sheet_name='Resumo Geral', index=False)
        
        # Write Análise por Fila
        all_filas = pd.DataFrame()
        for fila in ['Triagem', 'Emergência', 'Consulta', 'Internação', 'Alta']:
            if fila in data:
                fila_df = data[fila].copy()
                fila_df.insert(0, 'Fila', fila)
                all_filas = pd.concat([all_filas, fila_df])
        
        all_filas.to_excel(writer, sheet_name='Análise por Fila', index=False)
        
        # Format sheets
        workbook = writer.book
        for sheet_name in writer.sheets:
            format_excel(workbook, writer.sheets[sheet_name])

def create_comparison_excel(data, output_file):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write comparison tables
        for name, df in data.items():
            sheet_name = 'Comparação Original vs Melhorado V2' if 'MÉTRICAS GERAIS' in name else name
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format sheets
        workbook = writer.book
        for sheet_name in writer.sheets:
            format_excel(workbook, writer.sheets[sheet_name])

if __name__ == '__main__':
    # Process metrics data
    metrics_data = read_metrics_data('analise_metricas_v2.txt')
    create_metrics_excel(metrics_data, 'analise_metricas_v2.xlsx')
    
    # Process comparison data
    comparison_data = read_comparison_data('comparacao_modelos_v2.txt')
    create_comparison_excel(comparison_data, 'comparacao_modelos_v2.xlsx') 